# importing required modules
import re
import PyPDF2
from openpyxl import load_workbook
import os
from methods import follow, comment, ifperf, perf

# def follow(x, lis):
#     if (x in lis):
#         y = lis.index(x)
#         return lis[y + 1]
#     else:
#         return "Not in list"
#
# def comment(y, test1):
#     newstr = " "
#     for x in y:
#         if (test1.find(x) != -1):
#             count = test1.count(x)
#             i = 0
#             while (i < count):
#                 i += 1
#                 indx = test1.find(x)
#                 if(indx == -1 ):
#                     break
#                 test2 = test1[indx:]
#                 if len(re.findall(r"[\d]{1,2}/[\d]{1,2}/[\d]{4}", test2)) != 0:
#                     m = re.findall(r"[\d]{1,2}/[\d]{1,2}/[\d]{4}", test2)[0]
#                     z = test2.index(m) + len(m)
#                     newstr += test2[:z] + " "
#                     test1 = test1[0:indx] + test1[indx + z:]
#     return newstr
#
# def ifperf(y, test1, boox):
#     for x in y:
#         if (test1.find(x) != -1):
#                 indx = test1.find(x)
#                 if (indx == -1):
#                     break
#                 boox = 'True'
#     return boox
#
# def perf(y, test1):
#     newstr = " "
#     for x in y:
#         if (test1.find(x) != -1):
#             count = test1.count(x)
#             i = 0
#             while (i < count):
#                 i += 1
#                 indx = test1.find(x)
#                 if(indx == -1 ):
#                     break
#                 test2 = test1[indx:]
#                 if len(re.findall(r"[\d]{1,2}/[\d]{1,2}/[\d]{4}", test2)) != 0:
#                     m = re.findall(r"[\d]{1,2}/[\d]{1,2}/[\d]{4}", test2)[0]
#                 else:
#                     break
#                 z = test2.index(m) + len(m)
#                 newstr += test2[:z] + " "
#                 test1 = test1[0:indx] + test1[indx + z:]
#     return newstr

wb = load_workbook(filename=r"")
ws = wb.active

rootdir = r'C:\Users\Ephraim.Sun\Desktop\proj2\test1'

for subdir, dirs, files in os.walk(rootdir):
    for file in files:
        if file.endswith("CURRENT SCHEMATIC.pdf"):
            pdfFileObj = os.path.join(subdir, file)
            #pdfFileObj = open(r'C:\Users\Ephraim.Sun\Desktop\proj2\test\BASS 1\BASS 1 - CURRENT SCHEMATIC.pdf', 'rb')
            pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
            # print(pdfReader.numPages) # will give total number of pages in pdf

            # extract all texts from page1
            pageObj = pdfReader.getPage(0)
            page_content = pageObj.extractText()

            """get a list from the pdf"""
            list = page_content.splitlines()
            if(len(list) == 0):
                break
            print(list)

            API = follow('API / UWI', list)

            """take the contents from list and put the good stuff into list_content"""
            if ('al)' in list):
                start = list.index("al)", )
            # else:  # Bruh
            #     break
            if ('District' in list):
                end = list.index("District")
            elif ('0District' in list):
                end = list.index("0District")
            # else:  # Bruh
            #     break
            # list_content = []
            list_content = list[start + 1:end]
            # print(list_content)

            """Make all list_content into one big string test"""
            test = ''
            for i in list_content:
                x = i
                test = test + x
            #print(test)

            new_str = ''
            test1 = test
            boo = 'False'
            keywords = ["Hydraulic Fracture", "Hydraulic Frac", "Sand Frac", "Hydrl Frac-Acid Base", "Hydrl Frac",
                        "Acid Frac", "Fracture", "Frac"]

            new_str = comment(keywords, test1)

            perf_keywords = ["Perforated", "Perf", "perf"]
            perf1 = perf(perf_keywords, test1)
            #print(perf1)

            boo1 = ifperf(keywords, test1, boo)
            boo2 = ifperf(perf_keywords, test1, boo)
            if(boo1 == "False" or boo2 == "False"):
                boo3 = "False"
            else:
                boo3 = "True"

            row_cell = ' '
            isitperf = ' '
            perf_cell = ' '

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == API:
                        row_cell = cell.row
            # print(row_cell)

            # for row in ws.iter_rows():
            #     for cell in row:
            #         if cell.value == 'Comments':
            #             frac_cell = cell.coordinate

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == 'Feet of Perfs':
                        perf_cell = cell.coordinate

            # print(cell_Date)
            # if (row_cell != ' ' and frac_cell != ' '):
            #     # print(type(row_cell))
            #     # print(type(cell_Date))
            #     cell_col = frac_cell.rstrip(frac_cell[1])
            #     # print(cell_col)
            #     coordinates = str(cell_col) + str(row_cell)
            #     # print(coordinates)
            #     ws[coordinates] = new_str
            #     # print(DATE)

            if (row_cell != ' ' and perf_cell != ' '):
                # print(type(row_cell))
                # print(type(cell_Date))
                cell_col1 = perf_cell.rstrip(perf_cell[1])
                # print(cell_col)
                coordinates1 = str(cell_col1) + str(row_cell)
                # print(coordinates)
                ws[coordinates1] = perf1
                # print(DATE)

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == 'Perforated (y/n)':
                        isitperf = cell.coordinate

            if (row_cell != ' ' and perf_cell != ' '):
                if (boo3 == "True" and new_str != ' '):
                    # print(type(row_cell))
                    # print(type(cell_Date))
                    cell_col = isitperf.rstrip(isitperf[1])
                    #print(cell_col)
                    coordinates = str(cell_col) + str(row_cell)
                    # print(coordinates)
                    ws[coordinates] = "Y"
                    # print(DATE)

wb.save(filename=r'')