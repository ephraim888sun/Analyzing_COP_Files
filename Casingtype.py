import re
import PyPDF2
from openpyxl import load_workbook
import os
from methods import follow, casing

# def follow(x, lis):
#     if (x in lis):
#         y = lis.index(x)
#         return lis[y + 1]
#     else:
#         return "Not in list"
#
# def casing(y, test1):
#     newstr = " "
#     for x in y:
#         # if (test1.find(x) != -1):
#         #     newstr += x + " "
#         if (test1.find(x) != -1) and (x == "Open Hole"):
#             count = test1.count(x)
#             i = 0
#             while (i < count):
#                 i += 1
#                 indx = test1.find(x)
#                 if(indx == -1 ):
#                     break
#                 test2 = test1[indx:]
#                 m = re.findall(r"[\d]{1,2}/[\d]{1,2}/[\d]{4}", test2)[0]
#                 z = test2.index(m) + len(m)
#                 newstr += test2[:z] + " "
#                 test1 = test1[0:indx] + test1[indx + z:]
#     return newstr

wb = load_workbook(filename=r"")
ws = wb.active
rootdir = r'C:\Users\Ephraim.Sun\Desktop\proj2\test1'

for subdir, dirs, files in os.walk(rootdir):
    for file in files:
        if file.endswith("CURRENT SCHEMATIC.pdf"):
            pdfFileObj = os.path.join(subdir, file)
            #pdfFileObj = open(r'C:\Users\Ephraim.Sun\Desktop\proj2\test1\ALBRECHT 1\ALBRECHT 1 - CURRENT SCHEMATIC.pdf', 'rb')
            pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
            #print(pdfReader.numPages) # will give total number of pages in pdf

            #extract all texts from page1
            pageObj = pdfReader.getPage(0)
            page_content = pageObj.extractText()

            """get a list from the pdf"""
            list = page_content.splitlines()
            #print(list)

            API = follow('API / UWI', list)

            """Make all list into one big string test"""
            test = ''
            for i in list:
                x = i
                test = test + x
            #print(test)

            new_str = ''
            keywords = ["Open Hole", 'Cemented Liner', "Slotted Liner", "OH Packer", "Liner"]

            new_str = casing(keywords, test)

            row_cell = ' '
            cell_Date = ' '

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == API:
                        row_cell = cell.row
            # print(row_cell)

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == 'Casing Type':
                        cell_Date = cell.coordinate

            # print(cell_Date)
            if (row_cell != ' ' and cell_Date != ' '):
                # print(type(row_cell))
                # print(type(cell_Date))
                cell_col = cell_Date.rstrip(cell_Date[1])
                # print(cell_col)
                coordinates = str(cell_col) + str(row_cell)
                # print(coordinates)
                ws[coordinates] = new_str
                # print(DATE)

wb.save(filename=r'')