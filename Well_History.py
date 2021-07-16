import PyPDF2
from openpyxl import load_workbook
import os
from methods import follow

# def follow(x, lis):
#     if (x in lis):
#         y = lis.index(x)
#         return lis[y + 1]
#     else:
#         return "Not in list"

wb = load_workbook(filename=r"")
ws = wb.active
rootdir = r'C:\Users\Ephraim.Sun\Desktop\proj2\test1'


for subdir, dirs, files in os.walk(rootdir):
    for file in files:
        if file.endswith("Well History.pdf"):
            #pdfFileObj = open(r'C:\Users\Ephraim.Sun\Desktop\proj2\test1\ABADAN 1T\ABADAN 1T - Well History.pdf', 'rb')
            #pdfFileObj = open(r'C:\Users\Ephraim.Sun\Desktop\proj2\test1\ARLENE-A 2R\ARLENE-A 2R - Well History.pdf', 'rb')
            pdfFileObj = os.path.join(subdir, file)

            pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
            #print(pdfReader.numPages) # will give total number of pages in pdf

            #extract all texts from page1
            pageObj = pdfReader.getPage(0)
            page_content = pageObj.extractText()
            #print(page_content)
            """get a list from the pdf"""
            list = page_content.splitlines()
            #print(list)

            test = '' """string type"""
            for i in list:
                x = i
                test = test + x
            x = test.find('Zone: Zone')
            #print(x)
            if(x != -1):
                teststr = test[x:]
                y = teststr.find('Comment')
                newstr = teststr[6:y- 2]
            # print(newstr)

                row_cell = ' '
                cell_Date = ' '

                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == follow('API / UWI', list):
                            row_cell = cell.row
                        # print(row_cell)

                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == 'Zone Stimulated':
                            cell_Date = cell.coordinate
                # print(cell_Date)

                if(row_cell != ' ' and cell_Date != ' '):
                    # print(type(row_cell))
                    # print(type(cell_Date))
                    cell_col = cell_Date.rstrip(cell_Date[1])
                    # print(cell_col)
                    coordinates = str(cell_col) + str(row_cell)
                    # print(coordinates)
                    ws[coordinates] = newstr
                    # print(DATE)

wb.save(filename=r'')
