# Importing Required modules
import PyPDF2
from openpyxl import load_workbook
import os
from methods import follow, comment, frac_Date, open_hole

'''Open excel worksheet'''
wb = load_workbook(filename=r"C:\Users\Ephraim.Sun\Desktop\proj2\testrun.xlsx")
ws = wb.active

'''Loop through each file in rootdir'''
rootdir = r'C:\Users\Ephraim.Sun\Desktop\proj2\test'

for subdir, dirs, files in os.walk(rootdir):
    for file in files:
        if file.endswith("CURRENT SCHEMATIC.pdf"):
            pdfFileObj = os.path.join(subdir, file)
            pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

            '''extract all texts from page1'''
            pageObj = pdfReader.getPage(0)
            page_content = pageObj.extractText()

            '''get a list from the pdf'''
            list_pdf = page_content.splitlines()

            '''Get Well API #'''
            API = follow('API / UWI', list_pdf)

            '''Clean Up the List'''
            if 'al)' in list_pdf:
                start = list_pdf.index("al)", )
            if 'District' in list_pdf:
                end = list_pdf.index("District")
            elif '0District' in list_pdf:
                end = list_pdf.index("0District")
            list_content = list_pdf[start + 1:end]

            '''Make one big string test'''
            test = ''
            for i in list_content:
                x = i
                test = test + x

            '''Use Keywords to extract info from the string'''
            test1 = test
            keywords = ["Hydraulic Fracture", 'Hydraulic Frac-Other', "Hydraulic Frac", "Sand Frac", "Hydrl Frac-Acid Base", "Hydrl Frac",
                        "Acid Frac", "Fracture", "Frac"]
            new_str = comment(keywords, test1)
            frac_time = frac_Date(keywords, test1)
            frac_openhole = open_hole(keywords, test1)

            '''Find  rows and columns of each cell, and put extracted info in cell'''
            row_cell = ' '
            cell_Date = ' '
            cell_fracDate = ' '
            cell_openhole = ' '

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == API:
                        row_cell = cell.row

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == 'Comments':
                        cell_Date = cell.coordinate

            if row_cell != ' ' and cell_Date != ' ':
                cell_col = cell_Date.rstrip(cell_Date[1])
                coordinates = str(cell_col) + str(row_cell)
                ws[coordinates] = new_str

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == 'Frac Date':
                        cell_fracDate = cell.coordinate

            if row_cell != ' ' and cell_fracDate != ' ':
                cell_col1 = cell_fracDate.rstrip(cell_fracDate[1])
                coordinates1 = str(cell_col1) + str(row_cell)
                ws[coordinates1] = frac_time

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == 'Open Hole Feet':
                        cell_openhole = cell.coordinate

            if row_cell != ' ' and cell_openhole != ' ':
                cell_col2 = cell_openhole.rstrip(cell_openhole[1])
                coordinates2 = str(cell_col2) + str(row_cell)
                ws[coordinates2] = frac_openhole

'''Save excel worksheet'''
wb.save(filename=r'C:\Users\Ephraim.Sun\Desktop\proj2\testrun.xlsx')
