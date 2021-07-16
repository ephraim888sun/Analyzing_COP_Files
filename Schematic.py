# importing required modules
import PyPDF2
from openpyxl import load_workbook
import os
from methods import follow

# def follow(x, lis):
#     if (x in lis):
#         y = lis.index(x)
#         return lis[y + 1]
#     else:
#         return "Not in list"

wb = load_workbook(filename=r"")
ws = wb.active
rootdir = r'C:\Users\Ephraim.Sun\Desktop\proj2\test1'

for subdir, dirs, files in os.walk(rootdir):
    for file in files:
        if file.endswith("CURRENT SCHEMATIC.pdf"):
            pdfFileObj = os.path.join(subdir, file)
            #print(pdfFileObj)
            pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
            # print(pdfReader.numPages) # will give total number of pages in pdf

            # extract all texts from page1
            pageObj = pdfReader.getPage(0)
            page_content = pageObj.extractText()

            """get a list from the pdf"""
            list = page_content.splitlines()


            API = follow('API / UWI', list)
            # print(API)
            DATE = follow('Original Spud Date', list)
            # print(DATE)

            row_cell = ' '
            cell_Date = ' '

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == API:
                        row_cell = cell.row
            # print(row_cell)

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == 'Spud Date':
                        cell_Date = cell.coordinate

            # print(cell_Date)
            if(row_cell != ' ' and cell_Date != ' '):
                # print(type(row_cell))
                # print(type(cell_Date))
                cell_col = cell_Date.rstrip(cell_Date[1])
                # print(cell_col)
                coordinates = str(cell_col) + str(row_cell)
                # print(coordinates)
                ws[coordinates] = DATE
                # print(DATE)
wb.save(filename=r'')